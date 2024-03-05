import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

df = pd.read_excel(r'D:\Prabu developement\panda\starting\start.xlsx')

df[['Copy_Component 1', 'Copy_Component 2', 'Copy_Component 3',]] = df[['Component 1', 'Component 2', 'Component 3',]].copy()

columns_to_ffill = df.columns[(df.columns != 'Component 1') & (df.columns != 'Component 2') & (df.columns != 'Component 3') & (df.columns != 'Component 4')]

df[columns_to_ffill] = df[columns_to_ffill].ffill(axis=0)


column_to_exclude = ['Component 1', 'Component 2', 'Component 3','Part Number']
df_excluded_column = df.drop(column_to_exclude, axis=1)

df['depth'] = df_excluded_column.count(axis=1)

df["Hierarchy"] = df.apply(lambda row: "||".join(row[['Copy_Component 1', 'Copy_Component 2', 'Copy_Component 3', 'Component 4']].dropna().astype(str)), axis=1)

df.drop(['Copy_Component 1', 'Copy_Component 2', 'Copy_Component 3'], axis=1,inplace=True)

df.to_excel("completedfully2.xlsx",index=False)

wb = load_workbook("completedfully2.xlsx")
0
ws = wb.active

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = Border(left=Side(border_style="thin", color="000000"),
                             right=Side(border_style="thin", color="000000"),
                             top=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="000000"))

wb.save("completedfully_bordered.xlsx")